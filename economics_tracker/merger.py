from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from economics_tracker.config import EXPORT_COLUMNS
from economics_tracker.exporter import export_articles_to_excel
from economics_tracker.models import Article


def _load_articles_from_xlsx(path: Path) -> list[Article]:
    """Read all articles from both 'selected' and 'rejected' sheets."""
    try:
        wb = load_workbook(path)
    except Exception:
        return []

    articles: list[Article] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row <= 1:
            continue
        headers = [cell.value for cell in ws[1]]
        col = {name: idx for idx, name in enumerate(headers)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            def get(name: str) -> str:
                idx = col.get(name)
                return str(row[idx] or "") if idx is not None else ""

            authors_raw = get("authors")
            authors = [a.strip() for a in authors_raw.split(";") if a.strip()]
            jel_raw = get("jel_codes")
            jel = [j.strip() for j in jel_raw.split(";") if j.strip()]
            kw_raw = get("matched_keywords")
            keywords = [k.strip() for k in kw_raw.split(";") if k.strip()]

            article = Article(
                journal=get("journal"),
                title=get("title"),
                url=get("url"),
                date=get("date"),
                authors=authors,
                abstract=get("abstract"),
                doi=get("doi"),
                jel_codes=jel,
                matched_keywords=keywords,
                decision=get("decision"),
                reason=get("reason"),
                source=get("source"),
            )
            if article.title:
                articles.append(article)
    return articles


def _dedupe(articles: list[Article]) -> list[Article]:
    seen: set[tuple[str, str]] = set()
    unique: list[Article] = []
    for article in articles:
        key = article.dedupe_key()
        if key in seen:
            continue
        seen.add(key)
        unique.append(article)
    return unique


def merge_daily_files(input_paths: list[Path], output_path: Path) -> Path:
    """Merge multiple daily xlsx files into one deduplicated Excel."""
    all_articles: list[Article] = []
    for path in input_paths:
        all_articles.extend(_load_articles_from_xlsx(path))
    unique = _dedupe(all_articles)
    return export_articles_to_excel(unique, output_path)
