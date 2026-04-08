from pathlib import Path

import pytest

from economics_tracker.exporter import Workbook, export_articles_to_excel
from economics_tracker.models import Article


pytestmark = pytest.mark.skipif(Workbook is None, reason="openpyxl not installed")


def test_export_splits_selected_and_rejected(tmp_path: Path) -> None:
    output = tmp_path / "daily.xlsx"
    articles = [
        Article(
            journal="A",
            title="Selected",
            url="https://example.com/1",
            decision="selected",
            reason="priority JEL match: D",
        ),
        Article(
            journal="B",
            title="Rejected",
            url="https://example.com/2",
            decision="rejected",
            reason="no priority JEL or keyword signal",
        ),
    ]

    export_articles_to_excel(articles, output)

    from openpyxl import load_workbook

    workbook = load_workbook(output)
    selected = workbook["selected"]
    rejected = workbook["rejected"]

    assert selected.max_row == 2
    assert rejected.max_row == 2
    assert selected["B2"].value == "Selected"
    assert rejected["B2"].value == "Rejected"
